# -*- coding: utf-8 -*-
################################################################
import sys
import os
import pandas as pd
import sqlite3 as sq
import re
from openpyxl import load_workbook
################################################################
if sys.platform == 'linux': 
    Base=os.path.expanduser('~') + 'VKHCG'
else:
    Base='C:/VKHCG'
################################################################
print('################################')
print('Working Base :',Base, ' using ', sys.platform)
print('################################')
################################################################
sInputTemplateName='00-RawData/Balance-Sheet-Template.xlsx'
################################################################
sOutputFileName='06-Report/01-EDS/02-Python/Report-Balance-Sheet'
Company='04-Clark'
################################################################
sDatabaseName=Base + '/' + Company + '/06-Report/SQLite/clark.db'
conn = sq.connect(sDatabaseName)
#conn = sq.connect(':memory:')
################################################################
### Import Balance Sheet Data
################################################################
for y in range(1,13):
    sInputFileName='00-RawData/BalanceSheets' + str(y).zfill(2) + '.csv'
    sFileName=Base + '/' + Company + '/' + sInputFileName
    print('################################')
    print('Loading :',sFileName)
    print('################################')
    ForexDataRaw=pd.read_csv(sFileName,header=0,low_memory=False, encoding="latin-1")
    print('################################')
    ################################################################
    ForexDataRaw.index.names = ['RowID']
    sTable='BalanceSheets'
    print('Storing :',sDatabaseName,' Table:',sTable)
    if y == 1:
        print('Load Data')
        ForexDataRaw.to_sql(sTable, conn, if_exists="replace")
    else:
        print('Append Data')        
        ForexDataRaw.to_sql(sTable, conn, if_exists="append")
################################################################
sSQL="SELECT \
            Year, \
            Quarter, \
            Country, \
            Company, \
            CAST(Year AS INT) || 'Q' || CAST(Quarter AS INT) AS sDate, \
            Company || ' (' || Country || ')' AS sCompanyName , \
            CAST(Year AS INT) || 'Q' || CAST(Quarter AS INT) || '-' ||\
            Company || '-' || Country AS sCompanyFile \
        FROM BalanceSheets \
        GROUP BY \
            Year, \
            Quarter, \
            Country, \
            Company \
        HAVING Year is not null \
       ;"
sSQL=re.sub("\s\s+", " ", sSQL)
sDatesRaw=pd.read_sql_query(sSQL, conn)
print(sDatesRaw.shape)
sDates=sDatesRaw.head(5)
################################################################
## Loop Dates
################################################################
for i in range(sDates.shape[0]):
    sFileName=Base + '/' + Company + '/' + sInputTemplateName
    wb = load_workbook(sFileName)
    ws=wb.get_sheet_by_name("Balance-Sheet")
    sYear=sDates['sDate'][i]
    sCompany=sDates['sCompanyName'][i]
    sCompanyFile=sDates['sCompanyFile'][i]
    sCompanyFile=re.sub("\s+", "", sCompanyFile)
    
    ws['D3'] = sYear
    ws['D5'] = sCompany
    
    sFields = pd.DataFrame(
            [
           ['Cash','D16', 1],
           ['Accounts_Receivable','D17', 1],
           ['Doubtful_Accounts','D18', 1],
           ['Inventory','D19', 1],
           ['Temporary_Investment','D20', 1],
           ['Prepaid_Expenses','D21', 1],
           ['Long_Term_Investments','D24', 1],
           ['Land','D25', 1],
           ['Buildings','D26', 1],
           ['Depreciation_Buildings','D27', -1],
           ['Plant_Equipment','D28', 1],
           ['Depreciation_Plant_Equipment','D29', -1],
           ['Furniture_Fixtures','D30', 1],
           ['Depreciation_Furniture_Fixtures','D31', -1],
           ['Accounts_Payable','H16', 1],
           ['Short_Term_Notes','H17', 1],
           ['Current_Long_Term_Notes','H18', 1],
           ['Interest_Payable','H19', 1],
           ['Taxes_Payable','H20', 1],
           ['Accrued_Payroll','H21', 1],
           ['Mortgage','H24', 1],
           ['Other_Long_Term_Liabilities','H25', 1],
           ['Capital_Stock','H30', 1]
           ]
            )
    
    nYear=str(int(sDates['Year'][i]))
    nQuarter=str(int(sDates['Quarter'][i]))
    sCountry=str(sDates['Country'][i])
    sCompany=str(sDates['Company'][i])
    
    sFileName=Base + '/' + Company + '/' + sOutputFileName + \
    '-' + sCompanyFile + '.xlsx'
    
    print(sFileName)
        
    for j in range(sFields.shape[0]):
        
        sSumField=sFields[0][j]
        sCellField=sFields[1][j]
        nSumSign=sFields[2][j]
        
        sSQL="SELECT  \
               Year, \
               Quarter, \
               Country, \
               Company, \
               SUM(" + sSumField + ") AS nSumTotal \
            FROM BalanceSheets \
            GROUP BY \
               Year, \
               Quarter, \
               Country, \
               Company \
            HAVING \
                Year=" + nYear + " \
            AND \
                Quarter=" + nQuarter + " \
            AND \
                Country='" + sCountry + "' \
            AND \
                Company='" + sCompany + "' \
           ;"
        sSQL=re.sub("\s\s+", " ", sSQL)
        sSumRaw=pd.read_sql_query(sSQL, conn)
        
        ws[sCellField] = sSumRaw["nSumTotal"][0] * nSumSign
        
        print('Set cell',sCellField,' to ', sSumField,'Total')
        
    wb.save(sFileName)